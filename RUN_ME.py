#!/usr/bin/env python3
"""Run this script to generate the Excel analysis file"""
import subprocess
import sys
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# Install openpyxl if needed
subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--break-system-packages', '-q'], stderr=subprocess.DEVNULL)

# Import what we need
import random
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)

HEADER_FILL = PatternFill("solid", fgColor="6B7F5E")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="E8E4DE")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="6B7F5E")
SUBTITLE_FONT = Font(name="Arial", size=11, color="888888")
GOOD_FILL = PatternFill("solid", fgColor="E8F5E9")
WARN_FILL = PatternFill("solid", fgColor="FFF8E1")
BAD_FILL = PatternFill("solid", fgColor="FFEBEE")
BORDER = Border(bottom=Side(style="thin", color="E0E0E0"))

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_row(ws, row, max_col, fill=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BODY_FONT
        cell.border = BORDER
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

def auto_width(ws, max_col, min_w=10, max_w=40):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = min_w
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = max_len + 2

NAMES = ["Anna Andersson", "Erik Johansson", "Maria Karlsson", "Lars Nilsson", "Eva Eriksson", "Johan Larsson", "Karin Olsson", "Anders Persson", "Sara Svensson", "Per Gustafsson", "Birgitta Pettersson", "Magnus Jonsson", "Lena Jansson", "Henrik Hansson", "Elisabeth Bengtsson", "Olof Jönsson", "Ingrid Lindberg", "Gustav Lindström", "Margareta Jakobsson", "Karl Magnusson", "Christina Lindgren", "Fredrik Axelsson", "Ulla Bergström", "Sven Lundberg", "Helena Berglund", "Mikael Fredriksson", "Kristina Sandberg", "Jan Henriksson", "Malin Forsberg", "Nils Sjöberg", "Emma Wallin", "Daniel Engström", "Cecilia Eklund", "Peter Danielsson", "Monica Lundin", "Ola Håkansson", "Gunilla Björk", "Mattias Bergman", "Annika Mattsson", "Björn Fransson", "Sofia Lindqvist", "Thomas Samuelsson", "Katarina Nyström", "Bengt Holmberg", "Jenny Arvidsson", "Rickard Löfgren", "Berit Söderberg", "David Nyberg", "Hanna Blomqvist", "Martin Claesson", "Elin Mårtensson", "Stefan Nordström", "Marianne Lundgren", "Tomas Eliasson", "Josefin Björklund", "Robert Bergqvist", "Barbro Isaksson", "Axel Lindholm", "Linda Nordin", "Mats Nygren", "Therese Lund", "Claes Ström", "Susanne Sundberg", "Oscar Hermansson", "Victoria Åberg", "Göran Ekström", "Camilla Holmgren", "Andreas Hedlund", "Åsa Dahlberg", "Joakim Hellström", "Lisa Sjögren", "Patrik Abrahamsson", "Agneta Falk", "Jesper Martinsson", "Ulrika Öberg"]

RELATIONS = ["Barn", "Make/Maka", "Syskon", "Barnbarn", "Förälder", "Annan släkting", "Vän/Kontakt"]
RELATION_WEIGHTS = [35, 20, 15, 10, 5, 10, 5]

CITIES = ["Stockholm", "Göteborg", "Malmö", "Uppsala", "Linköping", "Örebro", "Västerås", "Norrköping"]
OCCUPATIONS = ["Pensionär", "Lärare", "Sjuksköterska", "Ingenjör", "Ekonom", "Egenföretagare", "Säljare", "Chef"]

FEATURES = {
    "Nödbroms (dag 1-7)": {"cat": "guide", "relevance": ["Barn", "Make/Maka"]},
    "Bouppteckning": {"cat": "dokument", "relevance": ["Barn", "Make/Maka"]},
    "Arvskifte": {"cat": "dokument", "relevance": ["Barn", "Make/Maka"]},
    "Checklistor": {"cat": "guide", "relevance": ["Barn", "Make/Maka", "Syskon"]},
    "Fullmakt-generator": {"cat": "dokument", "relevance": ["Barn", "Make/Maka"]},
    "Ordlista": {"cat": "guide", "relevance": ["Barn", "Make/Maka"]},
    "Tidslinje": {"cat": "verktyg", "relevance": ["Barn", "Make/Maka"]},
    "Begravningsplanering": {"cat": "verktyg", "relevance": ["Barn", "Make/Maka"]},
    "Delägare-portal": {"cat": "samarbete", "relevance": ["Barn", "Make/Maka"]},
    "Juridisk AI-hjälp": {"cat": "verktyg", "relevance": ["Barn", "Make/Maka"]},
    "Dokument-skanner": {"cat": "verktyg", "relevance": ["Barn", "Make/Maka"]},
    "Arvskalkylator": {"cat": "verktyg", "relevance": ["Barn", "Make/Maka"]},
}

PLANS = [
    {"name": "Gratis", "price": 0},
    {"name": "Standard", "price": 899},
    {"name": "Premium", "price": 1499},
]

MISSING = ["Direktchatt med jurist", "Video-guider", "Sökfunktion", "Dark mode", "Realtidssamarbete"]

profiles = []
for i in range(75):
    age = random.choices([random.randint(25, 35), random.randint(36, 50), random.randint(51, 65), random.randint(66, 85)], weights=[15, 30, 35, 20])[0]
    relation = random.choices(RELATIONS, weights=RELATION_WEIGHTS)[0]
    tech = random.randint(1, 5)
    complexity = random.choice(["Enkelt", "Medel", "Komplext"])
    heirs = random.randint(1, 7)

    used_features = {}
    for feat, info in FEATURES.items():
        if relation in info["relevance"] and random.random() < 0.4 + tech * 0.08:
            used_features[feat] = random.randint(2, 5)

    chosen_plan = random.choices(["Gratis", "Standard", "Premium"], weights=[25, 50, 25])[0]
    price_opinion = random.choice(["Rimligt pris", "För dyrt", "Värt det", "Bra pris"])

    nav_score = min(10, max(1, round(5 + tech + random.gauss(0, 1))))
    understand_score = min(10, max(1, round(5 + tech * 0.5 + random.gauss(0, 1))))
    design_score = min(10, max(1, round(6.5 + random.gauss(0, 1))))
    nps = min(10, max(0, round(5 + tech * 0.7 + random.gauss(0, 1.5))))

    profiles.append({
        "id": i + 1,
        "name": NAMES[i],
        "age": age,
        "city": random.choice(CITIES),
        "occupation": random.choice(OCCUPATIONS),
        "relation": relation,
        "tech_level": tech,
        "complexity": complexity,
        "heirs": heirs,
        "used_features": used_features,
        "chosen_plan": chosen_plan,
        "price_opinion": price_opinion,
        "nav_score": nav_score,
        "understand_score": understand_score,
        "design_score": design_score,
        "nps": nps,
        "would_recommend": "Ja" if nps >= 7 else ("Kanske" if nps >= 5 else "Nej"),
        "missing": random.sample(MISSING, random.randint(0, 2)),
    })

wb = Workbook()

# Sheet 1: Summary
ws = wb.active
ws.title = "Sammanfattning"
ws.sheet_properties.tabColor = "6B7F5E"
ws.merge_cells("A1:G1")
ws["A1"] = "SISTA RESAN — Användaranalys"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A2:G2")
ws["A2"] = f"75 simulerade användarprofiler · Genererad {datetime.now().strftime('%Y-%m-%d %H:%M')}"
ws["A2"].font = SUBTITLE_FONT

r = 4
for label, value in [
    ("NYCKELTAL", ""), ("Antal profiler", 75), ("Genomsnittlig ålder", round(sum(p["age"] for p in profiles) / 75, 1)),
    ("", ""), ("BETYG (1-10)", "Snitt"), ("Navigation", round(sum(p["nav_score"] for p in profiles) / 75, 1)),
    ("Förståelse", round(sum(p["understand_score"] for p in profiles) / 75, 1)), ("Design", round(sum(p["design_score"] for p in profiles) / 75, 1)),
    ("NPS-poäng", round(sum(p["nps"] for p in profiles) / 75, 1)), ("", ""),
    ("REKOMMENDATION", "Antal"), ("Ja", sum(1 for p in profiles if p["would_recommend"] == "Ja")),
    ("Kanske", sum(1 for p in profiles if p["would_recommend"] == "Kanske")), ("Nej", sum(1 for p in profiles if p["would_recommend"] == "Nej")),
]:
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=value)
    if str(label).isupper() and label:
        ws.cell(row=r, column=1).font = BOLD_FONT
        ws.cell(row=r, column=1).fill = SUBHEADER_FILL
    r += 1

auto_width(ws, 3)

# Sheet 2: Profiles
ws2 = wb.create_sheet("Profiler")
ws2.sheet_properties.tabColor = "4A90D9"
headers = ["#", "Namn", "Ålder", "Stad", "Yrke", "Relation", "Teknikvana", "Komplexitet", "Delägare", "Plan", "Pris"]
for col, h in enumerate(headers, 1):
    ws2.cell(row=1, column=col, value=h)
style_header(ws2, 1, len(headers))
for i, p in enumerate(profiles, 2):
    ws2.cell(row=i, column=1, value=p["id"])
    ws2.cell(row=i, column=2, value=p["name"])
    ws2.cell(row=i, column=3, value=p["age"])
    ws2.cell(row=i, column=4, value=p["city"])
    ws2.cell(row=i, column=5, value=p["occupation"])
    ws2.cell(row=i, column=6, value=p["relation"])
    ws2.cell(row=i, column=7, value=p["tech_level"])
    ws2.cell(row=i, column=8, value=p["complexity"])
    ws2.cell(row=i, column=9, value=p["heirs"])
    ws2.cell(row=i, column=10, value=p["chosen_plan"])
    ws2.cell(row=i, column=11, value=0 if p["chosen_plan"] == "Gratis" else (899 if p["chosen_plan"] == "Standard" else 1499))
auto_width(ws2, len(headers))

# Sheet 3: Ratings
ws3 = wb.create_sheet("Betyg")
ws3.sheet_properties.tabColor = "E8A838"
headers3 = ["#", "Namn", "Navigation", "Förståelse", "Design", "NPS", "Rekommenderar?", "Prisåsikt"]
for col, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=col, value=h)
style_header(ws3, 1, len(headers3))
for i, p in enumerate(profiles, 2):
    ws3.cell(row=i, column=1, value=p["id"])
    ws3.cell(row=i, column=2, value=p["name"])
    ws3.cell(row=i, column=3, value=p["nav_score"])
    ws3.cell(row=i, column=4, value=p["understand_score"])
    ws3.cell(row=i, column=5, value=p["design_score"])
    ws3.cell(row=i, column=6, value=p["nps"])
    ws3.cell(row=i, column=7, value=p["would_recommend"])
    ws3.cell(row=i, column=8, value=p["price_opinion"])
    fill = GOOD_FILL if p["nps"] >= 8 else (BAD_FILL if p["nps"] <= 4 else None)
    style_row(ws3, i, len(headers3), fill)
auto_width(ws3, len(headers3))

# Sheet 4: Features
ws4 = wb.create_sheet("Funktioner")
ws4.sheet_properties.tabColor = "9B59B6"
feat_list = list(FEATURES.keys())
headers4 = ["#", "Namn", "Relation"] + feat_list
for col, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=col, value=h)
style_header(ws4, 1, len(headers4))
for i, p in enumerate(profiles, 2):
    ws4.cell(row=i, column=1, value=p["id"])
    ws4.cell(row=i, column=2, value=p["name"])
    ws4.cell(row=i, column=3, value=p["relation"])
    for j, feat in enumerate(feat_list):
        rating = p["used_features"].get(feat, "")
        cell = ws4.cell(row=i, column=j+4, value=rating)
        if isinstance(rating, int) and rating >= 4:
            cell.fill = GOOD_FILL
auto_width(ws4, 3)

# Sheet 5: Pricing
ws5 = wb.create_sheet("Prisanalys")
ws5.sheet_properties.tabColor = "27AE60"
ws5.merge_cells("A1:D1")
ws5["A1"] = "PRISANALYS"
ws5["A1"].font = TITLE_FONT
r = 3
ws5.cell(row=r, column=1, value="Plan")
ws5.cell(row=r, column=2, value="Antal")
ws5.cell(row=r, column=3, value="Andel")
style_header(ws5, r, 3)
for plan in PLANS:
    r += 1
    count = sum(1 for p in profiles if p["chosen_plan"] == plan["name"])
    ws5.cell(row=r, column=1, value=plan["name"])
    ws5.cell(row=r, column=2, value=count)
    ws5.cell(row=r, column=3, value=f"{round(count/75*100)}%")
auto_width(ws5, 3)

# Save
output_path = "sistaresan_användaranalys_75_profiler.xlsx"
wb.save(output_path)
print(f"Excel-fil sparad: {output_path}")
print(f"5 flikar genererade med 75 användarprofiler")
print(f"Navigation: {round(sum(p['nav_score'] for p in profiles) / 75, 1)}/10")
print(f"Förståelse: {round(sum(p['understand_score'] for p in profiles) / 75, 1)}/10")
print(f"Design: {round(sum(p['design_score'] for p in profiles) / 75, 1)}/10")
print(f"NPS: {round(sum(p['nps'] for p in profiles) / 75, 1)}")
print(f"{round(sum(1 for p in profiles if p['would_recommend'] == 'Ja') / 75 * 100)}% skulle rekommendera")
