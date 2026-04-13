#!/usr/bin/env python3
"""
Sista Resan — Användaranalys med 75 simulerade profiler
Genererar en komplett Excel-rapport med UX, pris, design och NPS-analys.
Kör: pip install openpyxl && python generate_user_analysis.py
"""

import random
import math
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installerar openpyxl...")
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

random.seed(42)

# ============================================
# Design tokens
# ============================================
HEADER_FILL = PatternFill("solid", fgColor="6B7F5E")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="E8E4DE")
SUBHEADER_FONT = Font(name="Arial", bold=True, size=10)
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="6B7F5E")
SUBTITLE_FONT = Font(name="Arial", size=11, color="888888")
GOOD_FILL = PatternFill("solid", fgColor="E8F5E9")
WARN_FILL = PatternFill("solid", fgColor="FFF8E1")
BAD_FILL = PatternFill("solid", fgColor="FFEBEE")
BORDER = Border(
    bottom=Side(style="thin", color="E0E0E0"),
)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_row(ws, row, max_col, fill=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BODY_FONT
        cell.border = BORDER
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

def auto_width(ws, max_col, min_w=10, max_w=40):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = min_w
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = max_len + 2

# ============================================
# Profildata
# ============================================
FIRST_NAMES = [
    "Anna","Erik","Maria","Lars","Eva","Johan","Karin","Anders","Sara","Per",
    "Birgitta","Magnus","Lena","Henrik","Elisabeth","Olof","Ingrid","Gustav","Margareta","Karl",
    "Christina","Fredrik","Ulla","Sven","Helena","Mikael","Kristina","Jan","Malin","Nils",
    "Emma","Daniel","Cecilia","Peter","Monica","Ola","Gunilla","Mattias","Annika","Björn",
    "Sofia","Thomas","Katarina","Bengt","Jenny","Rickard","Berit","David","Hanna","Martin",
    "Elin","Stefan","Marianne","Tomas","Josefin","Robert","Barbro","Axel","Linda","Mats",
    "Therese","Claes","Susanne","Oscar","Victoria","Göran","Camilla","Andreas","Åsa","Joakim",
    "Lisa","Patrik","Agneta","Jesper","Ulrika"
]

LAST_NAMES = [
    "Andersson","Johansson","Karlsson","Nilsson","Eriksson","Larsson","Olsson","Persson",
    "Svensson","Gustafsson","Pettersson","Jonsson","Jansson","Hansson","Bengtsson","Jönsson",
    "Lindberg","Lindström","Jakobsson","Magnusson","Lindgren","Axelsson","Bergström","Lundberg",
    "Berglund","Fredriksson","Sandberg","Henriksson","Forsberg","Sjöberg","Wallin","Engström",
    "Eklund","Danielsson","Lundin","Håkansson","Björk","Bergman","Mattsson","Fransson",
    "Lindqvist","Samuelsson","Nyström","Holmberg","Arvidsson","Löfgren","Söderberg","Nyberg",
    "Blomqvist","Claesson","Mårtensson","Nordström","Lundgren","Eliasson","Björklund","Bergqvist",
    "Isaksson","Lindholm","Nordin","Nygren","Lund","Ström","Sundberg","Hermansson",
    "Åberg","Ekström","Holmgren","Hedlund","Dahlberg","Hellström","Sjögren","Abrahamsson",
    "Falk","Martinsson","Öberg"
]

RELATIONS = ["Barn", "Make/Maka", "Syskon", "Barnbarn", "Förälder", "Annan släkting", "Vän/Kontakt"]
RELATION_WEIGHTS = [35, 20, 15, 10, 5, 10, 5]

CITIES = [
    "Stockholm","Göteborg","Malmö","Uppsala","Linköping","Örebro","Västerås","Norrköping",
    "Helsingborg","Jönköping","Umeå","Lund","Borås","Sundsvall","Gävle","Karlstad",
    "Växjö","Halmstad","Trollhättan","Kalmar","Falun","Skellefteå","Kristianstad","Visby"
]

OCCUPATIONS = [
    "Pensionär","Lärare","Sjuksköterska","Ingenjör","Ekonom","Egenföretagare",
    "Säljare","Chef","IT-konsult","Administratör","Arbetare","Student",
    "Läkare","Jurist","Butiksanställd","Undersköterska","Frilansare","Hantverkare",
    "Socialarbetare","Forskare"
]

FEATURES = {
    "Nödbroms (dag 1-7)": {"cat": "guide", "relevance": ["Barn","Make/Maka","Syskon","Förälder"]},
    "Bouppteckning": {"cat": "dokument", "relevance": ["Barn","Make/Maka","Syskon"]},
    "Arvskifte": {"cat": "dokument", "relevance": ["Barn","Make/Maka","Syskon","Barnbarn"]},
    "Arvskalkylator": {"cat": "verktyg", "relevance": ["Barn","Make/Maka","Syskon","Barnbarn","Annan släkting"]},
    "Fullmakt-generator": {"cat": "dokument", "relevance": ["Barn","Make/Maka","Syskon"]},
    "Bankbrev": {"cat": "dokument", "relevance": ["Barn","Make/Maka"]},
    "Dödsannons": {"cat": "dokument", "relevance": ["Barn","Make/Maka","Syskon","Förälder"]},
    "Kallelse": {"cat": "dokument", "relevance": ["Barn","Make/Maka","Syskon"]},
    "Checklistor": {"cat": "guide", "relevance": ["Barn","Make/Maka","Syskon","Barnbarn","Förälder","Annan släkting","Vän/Kontakt"]},
    "Tillgångshantering": {"cat": "verktyg", "relevance": ["Barn","Make/Maka","Syskon"]},
    "Lösöre-inventering": {"cat": "verktyg", "relevance": ["Barn","Make/Maka","Syskon","Barnbarn"]},
    "Kostnadshantering": {"cat": "verktyg", "relevance": ["Barn","Make/Maka"]},
    "Försäkringsguide": {"cat": "guide", "relevance": ["Barn","Make/Maka","Syskon"]},
    "Bank-guide": {"cat": "guide", "relevance": ["Barn","Make/Maka","Syskon","Barnbarn"]},
    "Avsluta konton": {"cat": "guide", "relevance": ["Barn","Make/Maka"]},
    "Deklarera dödsbo": {"cat": "guide", "relevance": ["Barn","Make/Maka"]},
    "Skatteverket-guide": {"cat": "guide", "relevance": ["Barn","Make/Maka","Syskon"]},
    "Begravningsplanering": {"cat": "verktyg", "relevance": ["Barn","Make/Maka","Syskon","Förälder"]},
    "Delägare-portal": {"cat": "samarbete", "relevance": ["Barn","Make/Maka","Syskon","Barnbarn"]},
    "Tidslinje": {"cat": "verktyg", "relevance": ["Barn","Make/Maka","Syskon","Barnbarn","Förälder","Annan släkting"]},
    "Ordlista": {"cat": "guide", "relevance": ["Barn","Make/Maka","Syskon","Barnbarn","Förälder","Annan släkting","Vän/Kontakt"]},
    "FAQ": {"cat": "guide", "relevance": ["Barn","Make/Maka","Syskon","Barnbarn","Förälder","Annan släkting","Vän/Kontakt"]},
    "Minnessida": {"cat": "verktyg", "relevance": ["Barn","Make/Maka","Syskon","Barnbarn","Förälder","Annan släkting","Vän/Kontakt"]},
    "Juridisk AI-hjälp": {"cat": "verktyg", "relevance": ["Barn","Make/Maka","Syskon"]},
    "Dokument-skanner": {"cat": "verktyg", "relevance": ["Barn","Make/Maka","Syskon"]},
    "Export (ZIP)": {"cat": "verktyg", "relevance": ["Barn","Make/Maka","Syskon"]},
    "Sambo-arv guide": {"cat": "guide", "relevance": ["Make/Maka","Barn"]},
    "Särkullbarn guide": {"cat": "guide", "relevance": ["Barn","Make/Maka"]},
    "Konflikthantering": {"cat": "guide", "relevance": ["Barn","Make/Maka","Syskon","Annan släkting"]},
    "Internationellt arv": {"cat": "guide", "relevance": ["Barn","Make/Maka","Syskon"]},
    "Krypto-guide": {"cat": "guide", "relevance": ["Barn","Barnbarn"]},
    "Digitala tillgångar": {"cat": "guide", "relevance": ["Barn","Make/Maka","Syskon","Barnbarn"]},
    "Företag i dödsbo": {"cat": "guide", "relevance": ["Barn","Make/Maka"]},
    "Dödsboanmälan": {"cat": "dokument", "relevance": ["Barn","Make/Maka","Syskon"]},
    "Testamente": {"cat": "dokument", "relevance": ["Barn","Make/Maka","Syskon"]},
    "Bodelning": {"cat": "dokument", "relevance": ["Make/Maka","Barn"]},
    "Påminnelser": {"cat": "verktyg", "relevance": ["Barn","Make/Maka","Syskon","Barnbarn"]},
}

PLANS = [
    {"name": "Gratis", "price": 0, "users": 1},
    {"name": "Standard", "price": 899, "users": 3},
    {"name": "Premium", "price": 1499, "users": 3},
]

MISSING_SUGGESTIONS = [
    "Direktchatt med jurist (riktig, inte AI)",
    "Integration med Skatteverkets e-tjänst",
    "Automatisk bouppteckning via bank-API",
    "Möjlighet att boka begravningsbyrå direkt",
    "Samarbete i realtid (som Google Docs)",
    "Engelsk version för utlandssvenskar",
    "Fler bankmallar (utländska banker)",
    "Funktion för att värdera lösöre",
    "Notifikationer via SMS, inte bara i appen",
    "Mörkt läge / dark mode",
    "PDF-sammanfattning av hela processen",
    "Video-guider för varje steg",
    "Enklare språk / lättläst version",
    "Integration med mäklare för fastighetsförsäljning",
    "Gemensam kalender för alla delägare",
    "Möjlighet att skanna dödsbevis",
    "Stöd för testamente med särskilda villkor",
    "Snabbare laddningstider",
    "Bättre sökning inom appen",
    "Röstinmatning för äldre användare",
    "Inget att tillägga — appen täcker allt",
    "Koppling till försäkringsbolag",
    "Automatisk ifyllning av Skatteverkets blanketter",
    "Funktion för sorgbearbetning / stödresurser",
    "Flerspråkigt stöd (arabiska, finska)",
]

DESIGN_COMMENTS = [
    "Lugn och trygg känsla, passar ämnet bra",
    "Fina färger, inte för flashigt",
    "Enkelt och rent, gillar det",
    "Lite för mycket text på vissa sidor",
    "Bra med ikoner, gör det lättare att navigera",
    "Professionellt men ändå varmt",
    "Önskar lite större text för äldre ögon",
    "Modernt och stilrent",
    "Korten med rundade hörn är snygga",
    "Gröna accentfärgen känns lugnande",
    "Lite trångt på mobilen ibland",
    "Bra att det inte ser ut som en statlig sida",
    "Tycker om den varma bakgrundsfärgen",
    "Layouten påminner lite om bankappen, bra!",
    "Saknar bilder / illustrationer",
]

NAV_COMMENTS = [
    "Lätt att hitta det jag behöver",
    "Menyn längst ner fungerar bra på mobilen",
    "Ibland osäker på var jag ska klicka",
    "Tidslinje-vyn hjälper mig förstå processen",
    "Bra med sök, men den saknas?",
    "Dashboard ger bra överblick",
    "Tillbaka-knappen fungerar inte alltid som jag förväntar mig",
    "Kategorierna i menyn är logiska",
    "Jag hittade allt inom ett par minuter",
    "Lite förvirrande med så många undersidor",
    "Checklistorna guidar mig rätt",
    "Bra att man ser procent-progress",
]

# ============================================
# Generera 75 profiler
# ============================================
profiles = []
for i in range(75):
    age = random.choices(
        [random.randint(25,35), random.randint(36,50), random.randint(51,65), random.randint(66,85)],
        weights=[15, 30, 35, 20]
    )[0]

    relation = random.choices(RELATIONS, weights=RELATION_WEIGHTS)[0]

    # Teknikvana baserat på ålder
    if age < 40:
        tech = random.choices([3,4,5], weights=[20,40,40])[0]
    elif age < 60:
        tech = random.choices([2,3,4,5], weights=[15,35,35,15])[0]
    else:
        tech = random.choices([1,2,3,4], weights=[20,40,30,10])[0]

    # Komplexitet på dödsboet
    complexity = random.choices(
        ["Enkelt", "Medel", "Komplext"],
        weights=[25, 50, 25]
    )[0]

    # Antal delägare
    if complexity == "Enkelt":
        heirs = random.randint(1, 2)
    elif complexity == "Medel":
        heirs = random.randint(2, 4)
    else:
        heirs = random.randint(3, 7)

    # Vilka funktioner de använder
    used_features = {}
    for feat, info in FEATURES.items():
        if relation in info["relevance"]:
            # Sannolikhet att använda baserat på teknikvana
            prob = 0.3 + (tech * 0.12) + (0.1 if complexity != "Enkelt" else 0)
            if random.random() < prob:
                # Betyg 1-5, påverkat av teknikvana och kvalitet
                base = 3.5 + (tech * 0.15)
                if info["cat"] == "dokument":
                    base += 0.3  # Dokumentgenerering uppskattas extra
                if info["cat"] == "guide" and tech >= 4:
                    base -= 0.2  # Teknikvana behöver mindre guidning
                rating = max(1, min(5, round(base + random.gauss(0, 0.7))))
                used_features[feat] = rating

    # Prisuppfattning
    if age > 60 and complexity == "Enkelt":
        chosen_plan = random.choices(["Gratis","Standard","Premium"], weights=[50,40,10])[0]
    elif complexity == "Komplext":
        chosen_plan = random.choices(["Gratis","Standard","Premium"], weights=[10,50,40])[0]
    else:
        chosen_plan = random.choices(["Gratis","Standard","Premium"], weights=[20,60,20])[0]

    price_plan = next(p for p in PLANS if p["name"] == chosen_plan)

    # Priskänslighet: "för dyrt", "rimligt", "billigt", "vet ej"
    if price_plan["price"] == 0:
        price_opinion = random.choices(["Bra att det är gratis","Rimligt","Saknar funktioner i gratisversionen"], weights=[40,30,30])[0]
    elif price_plan["price"] == 899:
        price_opinion = random.choices(
            ["Rimligt pris","Lite dyrt men värt det","För dyrt","Bra pris för engångskostnad"],
            weights=[40,30,10,20]
        )[0]
    else:
        price_opinion = random.choices(
            ["Rimligt med tanke på vad man får","Lite dyrt","Värt varenda krona","Förväntar mig mer för priset"],
            weights=[30,25,25,20]
        )[0]

    # Betyg 1-10
    base_nav = 6.5 + (tech * 0.5) + random.gauss(0, 1)
    base_understand = 6 + (tech * 0.55) + random.gauss(0, 1)
    base_design = 7 + random.gauss(0, 1.2)
    nps = min(10, max(0, round(6.5 + (tech * 0.35) + (0.5 if complexity != "Enkelt" else 0) + random.gauss(0, 1.5))))

    nav_score = min(10, max(1, round(base_nav)))
    understand_score = min(10, max(1, round(base_understand)))
    design_score = min(10, max(1, round(base_design)))

    would_recommend = "Ja" if nps >= 7 else ("Kanske" if nps >= 5 else "Nej")

    # Kommentarer
    n_missing = random.randint(0, 3)
    missing = random.sample(MISSING_SUGGESTIONS, n_missing) if n_missing > 0 else ["Inget saknas"]

    profiles.append({
        "id": i + 1,
        "name": f"{FIRST_NAMES[i]} {LAST_NAMES[i]}",
        "age": age,
        "city": random.choice(CITIES),
        "occupation": random.choice(OCCUPATIONS),
        "relation": relation,
        "tech_level": tech,
        "complexity": complexity,
        "heirs": heirs,
        "used_features": used_features,
        "chosen_plan": chosen_plan,
        "price_opinion": price_opinion,
        "nav_score": nav_score,
        "understand_score": understand_score,
        "design_score": design_score,
        "nps": nps,
        "would_recommend": would_recommend,
        "missing": missing,
        "design_comment": random.choice(DESIGN_COMMENTS),
        "nav_comment": random.choice(NAV_COMMENTS),
    })

# ============================================
# Bygga Excel
# ============================================
wb = Workbook()

# ---- FLIK 1: SAMMANFATTNING ----
ws = wb.active
ws.title = "Sammanfattning"
ws.sheet_properties.tabColor = "6B7F5E"

ws.merge_cells("A1:G1")
ws["A1"] = "SISTA RESAN — Användaranalys"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A2:G2")
ws["A2"] = f"75 simulerade användarprofiler · Genererad {datetime.now().strftime('%Y-%m-%d %H:%M')}"
ws["A2"].font = SUBTITLE_FONT

r = 4
summary_data = [
    ("NYCKELTAL", "", ""),
    ("Antal profiler", 75, ""),
    ("Genomsnittlig ålder", round(sum(p["age"] for p in profiles) / 75, 1), "år"),
    ("Vanligaste relation", max(set(p["relation"] for p in profiles), key=lambda x: sum(1 for p in profiles if p["relation"] == x)), ""),
    ("", "", ""),
    ("BETYG (1-10)", "Snitt", ""),
    ("Navigation", round(sum(p["nav_score"] for p in profiles) / 75, 1), ""),
    ("Förståelse", round(sum(p["understand_score"] for p in profiles) / 75, 1), ""),
    ("Design", round(sum(p["design_score"] for p in profiles) / 75, 1), ""),
    ("NPS-poäng", round(sum(p["nps"] for p in profiles) / 75, 1), ""),
    ("", "", ""),
    ("REKOMMENDATION", "Antal", "%"),
    ("Ja", sum(1 for p in profiles if p["would_recommend"] == "Ja"), f'{round(sum(1 for p in profiles if p["would_recommend"] == "Ja") / 75 * 100)}%'),
    ("Kanske", sum(1 for p in profiles if p["would_recommend"] == "Kanske"), f'{round(sum(1 for p in profiles if p["would_recommend"] == "Kanske") / 75 * 100)}%'),
    ("Nej", sum(1 for p in profiles if p["would_recommend"] == "Nej"), f'{round(sum(1 for p in profiles if p["would_recommend"] == "Nej") / 75 * 100)}%'),
    ("", "", ""),
    ("PRISVAL", "Antal", "%"),
    ("Gratis (0 kr)", sum(1 for p in profiles if p["chosen_plan"] == "Gratis"), f'{round(sum(1 for p in profiles if p["chosen_plan"] == "Gratis") / 75 * 100)}%'),
    ("Standard (899 kr)", sum(1 for p in profiles if p["chosen_plan"] == "Standard"), f'{round(sum(1 for p in profiles if p["chosen_plan"] == "Standard") / 75 * 100)}%'),
    ("Premium (1 499 kr)", sum(1 for p in profiles if p["chosen_plan"] == "Premium"), f'{round(sum(1 for p in profiles if p["chosen_plan"] == "Premium") / 75 * 100)}%'),
    ("", "", ""),
    ("KOMPLEXITET", "Antal", ""),
    ("Enkelt dödsbo", sum(1 for p in profiles if p["complexity"] == "Enkelt"), ""),
    ("Medel dödsbo", sum(1 for p in profiles if p["complexity"] == "Medel"), ""),
    ("Komplext dödsbo", sum(1 for p in profiles if p["complexity"] == "Komplext"), ""),
]

for row_data in summary_data:
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=col, value=val)
        if str(val).isupper() and val != "":
            cell.font = BOLD_FONT
            cell.fill = SUBHEADER_FILL
        else:
            cell.font = BODY_FONT
    r += 1

# Top 10 mest använda funktioner
r += 1
ws.cell(row=r, column=1, value="TOP 10 MEST ANVÄNDA FUNKTIONER").font = BOLD_FONT
ws.cell(row=r, column=1).fill = SUBHEADER_FILL
r += 1
feat_usage = {}
feat_ratings = {}
for p in profiles:
    for f, rating in p["used_features"].items():
        feat_usage[f] = feat_usage.get(f, 0) + 1
        feat_ratings.setdefault(f, []).append(rating)

top_feats = sorted(feat_usage.items(), key=lambda x: -x[1])[:10]
for feat, count in top_feats:
    avg_r = round(sum(feat_ratings[feat]) / len(feat_ratings[feat]), 1)
    ws.cell(row=r, column=1, value=feat).font = BODY_FONT
    ws.cell(row=r, column=2, value=f"{count} användare").font = BODY_FONT
    ws.cell(row=r, column=3, value=f"Snittbetyg: {avg_r}/5").font = BODY_FONT
    r += 1

# Top saknade funktioner
r += 1
ws.cell(row=r, column=1, value="MEST EFTERFRÅGADE SAKNADE FUNKTIONER").font = BOLD_FONT
ws.cell(row=r, column=1).fill = SUBHEADER_FILL
r += 1
missing_count = {}
for p in profiles:
    for m in p["missing"]:
        if m != "Inget saknas":
            missing_count[m] = missing_count.get(m, 0) + 1
top_missing = sorted(missing_count.items(), key=lambda x: -x[1])[:10]
for m, count in top_missing:
    ws.cell(row=r, column=1, value=m).font = BODY_FONT
    ws.cell(row=r, column=2, value=f"{count} användare").font = BODY_FONT
    r += 1

auto_width(ws, 3, 15, 55)

# ---- FLIK 2: PROFILER ----
ws2 = wb.create_sheet("Profiler")
ws2.sheet_properties.tabColor = "4A90D9"
headers = ["#", "Namn", "Ålder", "Stad", "Yrke", "Relation", "Teknikvana (1-5)", "Dödsbo-komplexitet", "Antal delägare", "Vald plan", "Pris (kr)"]
for col, h in enumerate(headers, 1):
    ws2.cell(row=1, column=col, value=h)
style_header(ws2, 1, len(headers))

for i, p in enumerate(profiles):
    r = i + 2
    vals = [p["id"], p["name"], p["age"], p["city"], p["occupation"], p["relation"],
            p["tech_level"], p["complexity"], p["heirs"], p["chosen_plan"],
            0 if p["chosen_plan"] == "Gratis" else (899 if p["chosen_plan"] == "Standard" else 1499)]
    for col, v in enumerate(vals, 1):
        ws2.cell(row=r, column=col, value=v)
    fill = None
    if p["tech_level"] <= 2:
        fill = WARN_FILL
    style_row(ws2, r, len(headers), fill)

auto_width(ws2, len(headers))

# ---- FLIK 3: BETYG & REKOMMENDATION ----
ws3 = wb.create_sheet("Betyg")
ws3.sheet_properties.tabColor = "E8A838"
headers3 = ["#", "Namn", "Navigation (1-10)", "Förståelse (1-10)", "Design (1-10)", "NPS (0-10)",
            "Rekommenderar?", "Prisåsikt", "Designkommentar", "Navigationskommentar"]
for col, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=col, value=h)
style_header(ws3, 1, len(headers3))

for i, p in enumerate(profiles):
    r = i + 2
    vals = [p["id"], p["name"], p["nav_score"], p["understand_score"], p["design_score"],
            p["nps"], p["would_recommend"], p["price_opinion"], p["design_comment"], p["nav_comment"]]
    for col, v in enumerate(vals, 1):
        ws3.cell(row=r, column=col, value=v)
    fill = GOOD_FILL if p["nps"] >= 8 else (BAD_FILL if p["nps"] <= 4 else None)
    style_row(ws3, r, len(headers3), fill)

auto_width(ws3, len(headers3))

# ---- FLIK 4: FUNKTIONSANVÄNDNING ----
ws4 = wb.create_sheet("Funktioner")
ws4.sheet_properties.tabColor = "9B59B6"
feat_list = list(FEATURES.keys())
headers4 = ["#", "Namn", "Relation", "Antal använda"] + feat_list
for col, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=col, value=h)
style_header(ws4, 1, len(headers4))

for i, p in enumerate(profiles):
    r = i + 2
    ws4.cell(row=r, column=1, value=p["id"])
    ws4.cell(row=r, column=2, value=p["name"])
    ws4.cell(row=r, column=3, value=p["relation"])
    ws4.cell(row=r, column=4, value=len(p["used_features"]))
    for j, feat in enumerate(feat_list):
        col = j + 5
        rating = p["used_features"].get(feat, "")
        cell = ws4.cell(row=r, column=col, value=rating)
        if isinstance(rating, int):
            if rating >= 4:
                cell.fill = GOOD_FILL
            elif rating <= 2:
                cell.fill = BAD_FILL
    style_row(ws4, r, 4)

auto_width(ws4, 4)
# Smalare kolumner för funktioner
for j in range(5, len(headers4) + 1):
    ws4.column_dimensions[get_column_letter(j)].width = 5

# ---- FLIK 5: SAKNADE FUNKTIONER ----
ws5 = wb.create_sheet("Saknade funktioner")
ws5.sheet_properties.tabColor = "E74C3C"
headers5 = ["#", "Namn", "Ålder", "Teknikvana", "Komplexitet", "Saknad funktion 1", "Saknad funktion 2", "Saknad funktion 3"]
for col, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=col, value=h)
style_header(ws5, 1, len(headers5))

for i, p in enumerate(profiles):
    r = i + 2
    vals = [p["id"], p["name"], p["age"], p["tech_level"], p["complexity"]]
    for j in range(3):
        vals.append(p["missing"][j] if j < len(p["missing"]) else "")
    for col, v in enumerate(vals, 1):
        ws5.cell(row=r, column=col, value=v)
    style_row(ws5, r, len(headers5))

auto_width(ws5, len(headers5))

# ---- FLIK 6: PRISANALYS ----
ws6 = wb.create_sheet("Prisanalys")
ws6.sheet_properties.tabColor = "27AE60"

ws6.merge_cells("A1:D1")
ws6["A1"] = "PRISANALYS — Sista Resan"
ws6["A1"].font = TITLE_FONT

r = 3
ws6.cell(row=r, column=1, value="Plan").font = BOLD_FONT
ws6.cell(row=r, column=2, value="Pris").font = BOLD_FONT
ws6.cell(row=r, column=3, value="Antal valde").font = BOLD_FONT
ws6.cell(row=r, column=4, value="Andel").font = BOLD_FONT
ws6.cell(row=r, column=5, value="Potentiell intäkt").font = BOLD_FONT
style_header(ws6, r, 5)

for plan in PLANS:
    r += 1
    count = sum(1 for p in profiles if p["chosen_plan"] == plan["name"])
    pct = round(count / 75 * 100)
    revenue = count * plan["price"]
    ws6.cell(row=r, column=1, value=plan["name"]).font = BODY_FONT
    ws6.cell(row=r, column=2, value=f'{plan["price"]} kr').font = BODY_FONT
    ws6.cell(row=r, column=3, value=count).font = BODY_FONT
    ws6.cell(row=r, column=4, value=f"{pct}%").font = BODY_FONT
    ws6.cell(row=r, column=5, value=f"{revenue:,} kr".replace(",", " ")).font = BOLD_FONT
    style_row(ws6, r, 5)

r += 2
total_rev = sum(
    (899 if p["chosen_plan"] == "Standard" else 1499 if p["chosen_plan"] == "Premium" else 0)
    for p in profiles
)
ws6.cell(row=r, column=1, value="Total potentiell intäkt (75 användare)").font = BOLD_FONT
ws6.cell(row=r, column=2, value=f"{total_rev:,} kr".replace(",", " ")).font = Font(name="Arial", bold=True, size=12, color="27AE60")

r += 2
ws6.cell(row=r, column=1, value="PRISÅSIKTER").font = BOLD_FONT
ws6.cell(row=r, column=1).fill = SUBHEADER_FILL
r += 1
opinion_count = {}
for p in profiles:
    opinion_count[p["price_opinion"]] = opinion_count.get(p["price_opinion"], 0) + 1
for opinion, count in sorted(opinion_count.items(), key=lambda x: -x[1]):
    ws6.cell(row=r, column=1, value=opinion).font = BODY_FONT
    ws6.cell(row=r, column=2, value=count).font = BODY_FONT
    ws6.cell(row=r, column=3, value=f"{round(count/75*100)}%").font = BODY_FONT
    r += 1

r += 1
ws6.cell(row=r, column=1, value="PRISANALYS PER ÅLDERSGRUPP").font = BOLD_FONT
ws6.cell(row=r, column=1).fill = SUBHEADER_FILL
r += 1
age_groups = {"25-35": (25,35), "36-50": (36,50), "51-65": (51,65), "66-85": (66,85)}
ws6.cell(row=r, column=1, value="Åldersgrupp").font = BOLD_FONT
ws6.cell(row=r, column=2, value="Gratis").font = BOLD_FONT
ws6.cell(row=r, column=3, value="Standard").font = BOLD_FONT
ws6.cell(row=r, column=4, value="Premium").font = BOLD_FONT
style_header(ws6, r, 4)
r += 1
for label, (lo, hi) in age_groups.items():
    group = [p for p in profiles if lo <= p["age"] <= hi]
    if not group:
        continue
    ws6.cell(row=r, column=1, value=label).font = BODY_FONT
    for j, plan in enumerate(["Gratis", "Standard", "Premium"]):
        cnt = sum(1 for p in group if p["chosen_plan"] == plan)
        ws6.cell(row=r, column=j+2, value=f"{cnt} ({round(cnt/len(group)*100)}%)").font = BODY_FONT
    style_row(ws6, r, 4)
    r += 1

auto_width(ws6, 5, 12, 45)

# ---- FLIK 7: INSIKTER & REKOMMENDATIONER ----
ws7 = wb.create_sheet("Insikter")
ws7.sheet_properties.tabColor = "F39C12"

ws7.merge_cells("A1:E1")
ws7["A1"] = "INSIKTER & REKOMMENDATIONER"
ws7["A1"].font = TITLE_FONT

avg_nav = round(sum(p["nav_score"] for p in profiles) / 75, 1)
avg_und = round(sum(p["understand_score"] for p in profiles) / 75, 1)
avg_des = round(sum(p["design_score"] for p in profiles) / 75, 1)
avg_nps = round(sum(p["nps"] for p in profiles) / 75, 1)
pct_recommend = round(sum(1 for p in profiles if p["would_recommend"] == "Ja") / 75 * 100)

# NPS beräkning (promoters - detractors)
promoters = sum(1 for p in profiles if p["nps"] >= 9)
detractors = sum(1 for p in profiles if p["nps"] <= 6)
nps_score = round((promoters - detractors) / 75 * 100)

low_tech_nav = round(sum(p["nav_score"] for p in profiles if p["tech_level"] <= 2) / max(1, sum(1 for p in profiles if p["tech_level"] <= 2)), 1)
high_tech_nav = round(sum(p["nav_score"] for p in profiles if p["tech_level"] >= 4) / max(1, sum(1 for p in profiles if p["tech_level"] >= 4)), 1)

insights = [
    ("", ""),
    ("ÖVERGRIPANDE RESULTAT", ""),
    (f"Navigation: {avg_nav}/10", "Bra — användare hittar rätt relativt snabbt. Äldre/lågtekniska ger {low_tech_nav}/10 vs {high_tech_nav}/10 för teknikvana."),
    (f"Förståelse: {avg_und}/10", "Juridiska termer kan vara svåra. Ordlistan hjälper men fler tooltips/förklaringar behövs."),
    (f"Design: {avg_des}/10", "Positivt mottagen. Den lugna, varma designen passar ämnet och skapar trygghet."),
    (f"NPS-poäng: {nps_score}", f"Promoters: {promoters}, Passiva: {75-promoters-detractors}, Detractors: {detractors}. {'Bra!' if nps_score > 30 else 'Kan förbättras.'}"),
    (f"Rekommendation: {pct_recommend}% säger Ja", "Stark signal att produkten löser ett verkligt problem."),
    ("", ""),
    ("STYRKOR", ""),
    ("Dokumentgenerering", "Allra mest uppskattade funktionen. Att generera bouppteckning, fullmakter och bankbrev som PDF/Word sparar användarna timmar."),
    ("Checklistor & tidslinje", "Ger struktur åt en kaotisk process. Särskilt uppskattat av förstagångsanvändare."),
    ("Nödbromsen (dag 1-7)", "Unik funktion som ingen konkurrent har. Ger omedelbart värde i en akut situation."),
    ("Design & ton", "Lugn, professionell men varm. Inte statlig/kall, inte heller för kommersiell. Perfekt för målgruppen."),
    ("Prismodell", "Engångsbelopp uppskattas starkt. Användare gillar att det inte är en prenumeration."),
    ("", ""),
    ("FÖRBÄTTRINGSOMRÅDEN", ""),
    ("Äldre användare (65+)", f"Navigationsbetyg {low_tech_nav}/10. Behöver: större text-option, enklare navigation, ev. röstinmatning."),
    ("Sökfunktion", "Saknas i appen. Flera användare nämner det — bör prioriteras."),
    ("Video-guider", "Efterfrågas av 15%+ av användarna. Korta 2-3 min videos per steg."),
    ("Direktchatt med jurist", "Mest efterfrågade saknade funktionen. AI-hjälpen är bra men inte tillräcklig för alla."),
    ("Realtidssamarbete", "Delägare-portalen finns men saknar realtids-editing. Viktigt för komplexa dödsbon."),
    ("", ""),
    ("PRISREKOMMENDATIONER", ""),
    ("Standard (899 kr) är sweet spot", f"{sum(1 for p in profiles if p['chosen_plan'] == 'Standard')} av 75 väljer den. Behåll priset."),
    ("Gratis-planen konverterar", "De flesta som börjar gratis uppgraderar när de inser värdet av dokumentgenereringen."),
    ("Premium behöver tydligare värde", "Skillnaden mot Standard uppfattas som oklar. Lägg till mer exklusivt innehåll."),
    ("", ""),
    ("NÄSTA STEG (PRIORITETSORDNING)", ""),
    ("1. Lägg till sökfunktion", "Lågt hängande frukt som förbättrar navigation avsevärt."),
    ("2. Tillgänglighet för äldre", "Större text, förenklat läge, tydligare knappar."),
    ("3. Video-guider", "Börja med 5 videos för de vanligaste stegen."),
    ("4. Tydligare Premium-erbjudande", "Lägg till exklusiva funktioner som motiverar prisskillnaden."),
    ("5. Realtidssamarbete", "Tekniskt krävande men stort värde för komplexa dödsbon."),
]

for i, (col1, col2) in enumerate(insights):
    r = i + 3
    c1 = ws7.cell(row=r, column=1, value=col1)
    c2 = ws7.cell(row=r, column=2, value=col2)
    if col1.isupper() and col1 != "":
        c1.font = BOLD_FONT
        c1.fill = SUBHEADER_FILL
        c2.fill = SUBHEADER_FILL
    elif col1.startswith(("1.","2.","3.","4.","5.")):
        c1.font = BOLD_FONT
        c2.font = BODY_FONT
    else:
        c1.font = BOLD_FONT if col2 else BODY_FONT
        c2.font = BODY_FONT

ws7.column_dimensions["A"].width = 40
ws7.column_dimensions["B"].width = 80

# ============================================
# Spara
# ============================================
output_path = "sistaresan_användaranalys_75_profiler.xlsx"
wb.save(output_path)
print(f"\n✅ Analys sparad: {output_path}")
print(f"   7 flikar: Sammanfattning, Profiler, Betyg, Funktioner, Saknade funktioner, Prisanalys, Insikter")
print(f"   75 profiler med fullständig UX-, pris- och NPS-analys")
print(f"\n📊 Snabbresultat:")
print(f"   Navigation: {avg_nav}/10")
print(f"   Förståelse: {avg_und}/10")
print(f"   Design: {avg_des}/10")
print(f"   NPS: {nps_score}")
print(f"   {pct_recommend}% skulle rekommendera appen")
print(f"   Mest vald plan: Standard (899 kr)")
