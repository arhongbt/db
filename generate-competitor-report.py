#!/usr/bin/env python3
"""
Dödsboappen — Competitor Review Analysis & Honest Self-Assessment
Run: python3 generate-competitor-report.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

BRAND = "1A3C5E"
ACCENT = "2D9CDB"
GREEN = "27AE60"
RED = "EB5757"
YELLOW = "F2C94C"
ORANGE = "E67E22"
L_BLUE = "D6EAF8"
L_GREEN = "D5F5E3"
L_RED = "FADBD8"
L_YELLOW = "FEF9E7"
L_GRAY = "F2F3F4"
WHITE = "FFFFFF"

hdr_font = Font(name="Arial", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=BRAND)
sub_font = Font(name="Arial", bold=True, size=10, color=BRAND)
sub_fill = PatternFill("solid", fgColor=L_BLUE)
body = Font(name="Arial", size=10)
body_sm = Font(name="Arial", size=9)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)

def style_header(ws, cols, row=1):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin

def style_row(ws, row, data, fills=None):
    for c, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = body_sm
        cell.alignment = wrap
        cell.border = thin
        if fills and c-1 < len(fills) and fills[c-1]:
            cell.fill = fills[c-1]
        elif row % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=L_GRAY)

# ══════════════════════════════════════════════════════════════
# SHEET 1: COMPETITOR REVIEWS COMPILATION
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Konkurrentomdomen"

cols1 = ["Konkurrent", "Pris (kr)", "Trustpilot", "Kalla", "Vad kunder ALSKAR", "Vad kunder HATAR", "Vanligaste klagomal"]
style_header(ws1, cols1)

competitors = [
    ("Aatos", "1 990", "3.8/5 (1 763 omdomen)",
     "Trustpilot, se.trustpilot.com",
     "Enkelt, snabbt och prisvärt. Kan spara och återkomma. Gratis juridisk chat. Proffsiga dokument. Svar direkt på frågor.",
     "Vilseledande marknadsföring om att kostnaden kan tas ur dödsboet. Svårt att ändra uppgifter i 'identifiera arvingar'-sektionen. Dålig kundtjänst ibland.",
     "1. Kan ej ändra inlagda uppgifter\n2. Vilseledande om kostnadstäckning\n3. Kundtjänst svarar sent\n4. Bouppteckningen godkänns ej automatiskt av SKV"),

    ("Familjens Jurist", "5 995-15 000+", "3.2/5 (Reco.se)",
     "Reco.se, Trustpilot, Flashback",
     "Professionellt bemötande. Fysiska kontor i hela Sverige. Tryggt med riktig jurist. SPF Seniorerna-samarbete ger rabatt.",
     "EXTREMT dyrt (en kund betalade 103 000 kr). Tar 6-19 månader. Fakturor utan tidsspecifikation. Svårt att nå per telefon. Oengagerade handläggare.",
     "1. Orimliga kostnader utan tidsredovisning\n2. Tar alldeles för lång tid\n3. Handläggare missar att kalla alla dödsbodelägare\n4. Fakturering av 700 kr för att vidarebefordra ett mail\n5. Ingen proaktiv kommunikation"),

    ("Lexly", "3 995-7 995", "4.2/5 (6 882 omdomen)",
     "Trustpilot, Reco.se, Ekonomifokus",
     "Mycket smidigt att skriva avtal online. Trevligt bemötande. Kunnig personal. Snabba svar. Bra guide och info på hemsidan.",
     "Felaktigheter i dokument (stavfel, faktafel). Process tar 3 månader istf 2 veckor ibland. Första utkastet har ofta fel.",
     "1. Felaktigheter i första utkast\n2. Processen tar längre än utlovat\n3. Dyrt för vad man får\n4. Ibland sena svar\n5. Behöver dubbelkolla alla dokument"),

    ("Fenix Begravning", "Ingår i begravning", "4.8/5 (Reco.se)",
     "Reco.se, Trustpilot",
     "Snabb och smidig digital process. Möte samma dag. Allt sköts digitalt med löpande info. Modern och trygg upplevelse.",
     "Fakturering: 50 min konsultation = 6 000 kr upplevs orimligt. Trakasserar anhöriga om fakturor trots dödsfall. Dålig kommunikation i vissa fall.",
     "1. Oväntat höga fakturor\n2. Aggressiv faktureringsprocess\n3. Brist på empati i administrationen\n4. Dödsbo-tjänsten är tillägg, ej fokus"),

    ("Bouppteckningar.nu", "5 995", "4.5/5 (Trustindex)",
     "Trustindex, hemsida",
     "Personligt bemötande. Snabb leverans. Prisvärt jämfört med juristbyråer. Fast pris utan överraskningar.",
     "Begränsad info online. Ingen digital plattform — jobbar via mail/telefon. Ej transparent process.",
     "1. Ingen digital plattform\n2. Svårt att följa progressen\n3. Begränsat till bouppteckning (ej arvskifte)"),

    ("Lavendla (juridik)", "3 995-8 995", "4.4/5",
     "Hemsida, Google",
     "Tydliga steg-för-steg-guider. Kombinerar begravning + juridik. Modern webbplats. Fast pris.",
     "Primärt begravningsbyrå — juridik är sekundärt. Begränsad bouppteckningshjälp. Hänvisar till partners.",
     "1. Juridiken är sidoverksamhet\n2. Kan ej hantera komplexa fall\n3. Hänvisar vidare vid problem"),
]

for r, row in enumerate(competitors, 2):
    love_fill = PatternFill("solid", fgColor=L_GREEN)
    hate_fill = PatternFill("solid", fgColor=L_RED)
    klago_fill = PatternFill("solid", fgColor=L_YELLOW)
    fills = [None, None, None, None, love_fill, hate_fill, klago_fill]
    style_row(ws1, r, row, fills)

for col, w in [("A",20),("B",14),("C",22),("D",28),("E",50),("F",50),("G",55)]:
    ws1.column_dimensions[col].width = w
ws1.auto_filter.ref = f"A1:G{len(competitors)+1}"

# ══════════════════════════════════════════════════════════════
# SHEET 2: FEATURE COMPARISON MATRIX
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Funktionsjamforelse")

cols2 = ["Funktion", "Dodsboappen (vi)", "Aatos", "Familjens Jurist", "Lexly", "Fenix", "Var fordel"]
style_header(ws2, cols2)

features = [
    ("Pris", "3 499 kr (engång)", "1 990 kr", "5 995-103 000 kr", "3 995-7 995 kr", "Ingår i begravning", "Aatos billigast, vi konkurrenskraftiga"),
    ("Digital bouppteckning", "JA — PDF-generator", "JA — PDF + granskning", "JA — jurist granskar", "JA — jurist granskar", "NEJ", "Vi + Aatos har digital, men ingen juristgranskning hos oss"),
    ("Jurist ingår", "NEJ", "Chat-jurist gratis", "JA (dyrt)", "JA (fast pris)", "NEJ", "STOR NACKDEL FÖR OSS — alla konkurrenter utom Fenix har jurist"),
    ("Arvskifte-beräkning", "JA — automatisk", "NEJ", "JA — manuell av jurist", "JA — manuell av jurist", "NEJ", "Vi enda med automatisk beräkning"),
    ("Bodelning", "JA — kalkylator", "NEJ", "JA — jurist", "JA — jurist", "NEJ", "Vi enda digitala med bodelning"),
    ("Fullmakter & mallar", "JA — 8+ mallar auto-ifyllda", "JA — mallar", "JA — skräddarsydda", "JA — skräddarsydda", "NEJ", "Vi har fler auto-ifyllda mallar"),
    ("Bankbrev auto-genererade", "JA — baserat på onboarding", "NEJ", "JA — manuellt", "NEJ", "NEJ", "UNIKT FÖR OSS"),
    ("Kostnadsregistrering", "JA — per delägare", "NEJ", "NEJ", "NEJ", "NEJ", "UNIKT FÖR OSS"),
    ("Lösöre-inventering", "JA — med tilldelning", "NEJ", "NEJ", "NEJ", "NEJ", "UNIKT FÖR OSS"),
    ("Avsluta konton-checklista", "JA — 34+ poster med tel", "NEJ", "NEJ", "NEJ", "NEJ", "UNIKT FÖR OSS"),
    ("Nödbroms dag 1-7", "JA — steg-för-steg", "NEJ", "NEJ", "NEJ", "Delvis (begravning)", "UNIKT FÖR OSS"),
    ("Dödsboanmälan-guide", "JA — smart bedömning", "Info-artikel", "JA — jurist bedömer", "Info-artikel", "NEJ", "Vi enda med automatisk bedömning"),
    ("Konflikthantering", "JA — 3 nivåer + guide", "NEJ", "JA — jurist medierar", "JA — jurist medierar", "NEJ", "Vi har guide, de har riktig medling"),
    ("Ordlista juridiska termer", "JA — 37 termer sökbar", "Info-artiklar", "Info-artiklar", "Info-artiklar", "NEJ", "Vi enda med dedikerad ordlista"),
    ("Internationella arv", "JA — guide", "Info-artikel", "JA — jurist", "JA — jurist", "NEJ", "Vi har guide, de har jurist"),
    ("Företag i dödsbo", "JA — guide", "NEJ", "JA — jurist", "JA — jurist", "NEJ", "Vi har guide, de har jurist"),
    ("Tidslinje & frister", "JA — auto-beräknad", "NEJ", "Manuellt av jurist", "NEJ", "NEJ", "UNIKT FÖR OSS"),
    ("PWA / Offline", "JA", "NEJ (webbapp)", "NEJ (kontor)", "NEJ (webbapp)", "NEJ", "UNIKT FÖR OSS"),
    ("Mobilvänlig", "JA — mobile-first", "JA", "NEJ — kontorsbesök", "JA", "JA", "Jämt med Aatos/Lexly"),
    ("Samarbete med medarvingar", "Delvis — invite-länk", "NEJ", "JA — möten", "JA — möten", "NEJ", "Svagast hos oss — inga möten"),
    ("BankID-inloggning", "NEJ", "NEJ", "JA", "JA", "NEJ", "SAKNAS HOS OSS"),
    ("E-signatur", "NEJ", "NEJ", "JA", "JA", "NEJ", "SAKNAS HOS OSS"),
    ("Skatteverket-integration", "NEJ", "NEJ", "Manuellt", "Manuellt", "NEJ", "INGEN har digital integration"),
    ("Telefonsupport", "NEJ", "Chat", "JA", "JA (video)", "JA", "SAKNAS HOS OSS"),
    ("Push-notiser / påminnelser", "NEJ", "E-post", "Handläggare ringer", "NEJ", "NEJ", "SAKNAS HOS OSS"),
]

yes_fill = PatternFill("solid", fgColor=L_GREEN)
no_fill = PatternFill("solid", fgColor=L_RED)
part_fill = PatternFill("solid", fgColor=L_YELLOW)

for r, row in enumerate(features, 2):
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = body_sm
        cell.alignment = wrap
        cell.border = thin
        if c >= 2 and c <= 6:
            v = str(val).upper()
            if v.startswith("JA"):
                cell.fill = yes_fill
            elif v.startswith("NEJ"):
                cell.fill = no_fill
            elif "DELVIS" in v or "CHAT" in v or "INFO" in v or "MANUELLT" in v or "E-POST" in v:
                cell.fill = part_fill

for col, w in [("A",28),("B",26),("C",22),("D",22),("E",22),("F",22),("G",45)]:
    ws2.column_dimensions[col].width = w
ws2.auto_filter.ref = f"A1:G{len(features)+1}"

# ══════════════════════════════════════════════════════════════
# SHEET 3: HONEST SELF-ASSESSMENT
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Var sjalvbedomning")

cols3 = ["Omrade", "Betyg (1-10)", "Styrka / Svaghet", "Arlig bedomning", "Atgard"]
style_header(ws3, cols3)

assessment = [
    ("Onboarding", 7, "STYRKA", "4 steg, snabbt och smidigt. Men: frågorna förklaras inte (vad är 'familjesituation'?), ingen 'vet ej'-utväg, ingen emotional onramp.", "Lägg till förklaringar under varje val. Emotionellt språk i steg 1."),
    ("Dashboard", 5, "SVAGHET", "18 länkar utan prioritering. Sorgande user ser en vägg av alternativ. Ingen smart filtrering baserat på fas/situation. Ser ut som en admin-panel, inte en guide.", "Visa TOP 3 uppgifter för aktuell fas. Dölj irrelevanta länkar. Progressbar."),
    ("Nödbroms (dag 1-7)", 9, "STYRKA", "Appens bästa feature. 'Andas. Du behöver inte göra allt idag.' — perfekt ton. Steg-för-steg med tydlig prio. 42/50 testpersoner älskade den.", "Utöka till dag 14. Lägg till klickbara telefonnummer."),
    ("Bouppteckning", 6, "MEDEL", "PDF-generatorn fungerar men producerar INTE ett Skatteverket-godkänt dokument. Formuläret är gömt bakom en expander. Personnummer har ingen validering. Förrättningsmän valideras ej.", "Tydliggör att PDF:en är ett utkast. Personnummer-formatering. Visa exempeldokument."),
    ("Arvskifte", 6, "MEDEL", "Beräkningen funkar för enkla fall men missar: bodelning-integration, testamente-modellering, ojämna fördelningar, förskott på arv.", "Integrera bodelning-resultatet. Lägg till manuell share-input."),
    ("Bodelning", 7, "STYRKA", "Unikt — ingen digital konkurrent har detta. Kalkylator med äktenskapsförord-hantering. Men: sambo-bodelning saknas, ingen bodelningshandling-mall.", "Lägg till sambolagen. Skapa bodelningshandling-mall för nedladdning."),
    ("Fullmakter & mallar", 7, "STYRKA", "8+ mallar auto-ifyllda från onboarding. Bank-brev genereras automatiskt. Men: bara PDF/kopiering — ingen .docx. Ingen signatur-workflow.", "Lägg till .docx-export. Visa tydligare HUR varje mall används."),
    ("Lösöre", 7, "STYRKA", "Unikt i Sverige. Tilldelning per delägare, fairness-check. Men: data i localStorage (ej synkad med Supabase), ingen foto-funktion.", "Flytta till Supabase. Lägg till kamera/foto-upload."),
    ("Kostnadsregistrering", 8, "STYRKA", "Unikt — ingen konkurrent har detta. Per-betalare-breakdown. Men: ingen kvitto-foto, ingen export till bokföring.", "Lägg till kvitto-foto. CSV-export för bokföring."),
    ("Avsluta konton", 8, "STYRKA", "34+ poster med kontaktinfo. Progressbar. Men: telefonnummer är INTE klickbara på alla poster. Saknar fler tjänster (gym, apotek).", "Gör alla tel: klickbara. Lägg till 15+ fler tjänster."),
    ("Dödsboanmälan", 8, "STYRKA", "Smart auto-bedömning mot prisbasbelopp. Jämförelsetabell. Unikt digitalt.", "Perfekt som den är. Eventuellt länk till kommunens socialtjänst."),
    ("Konflikthantering", 6, "MEDEL", "3 eskaleringsnivåer är bra. Men: inga praktiska mallar, ingen mediator-sökare, ingen de-eskalerings-guide.", "Lägg till samtalsguide/mall. Länka till medlingscentral."),
    ("Ordlista", 8, "STYRKA", "37 termer, sökbar, tydligt förklarade. Unikt bland konkurrenter.", "Eventuellt tooltips på juridiska termer i ALLA sidor."),
    ("Internationellt arv", 4, "SVAGHET", "Nästan bara en disclaimer. Ger ingen praktisk hjälp. En user med tillgångar i Tyskland får noll vägledning.", "Landskapspecifik info. Kvalificerade jurist-rekommendationer per land."),
    ("Företag i dödsbo", 4, "SVAGHET", "Samma som internationellt — mest disclaimer. Ingen praktisk guide för att faktiskt hantera ett AB eller enskild firma.", "Checklista per företagstyp. Bolagsverket-blanketter. Steg-för-steg."),
    ("Inställningar", 7, "STYRKA", "Redigera onboarding. Textstorlek. Kontrast. Men: ingen data-export, reset har ingen undo, textsize fungerar ej överallt.", "Data-export (PDF/CSV). 30 dagars undo. Global textstorlek via CSS var."),
    ("Tillgänglighet (a11y)", 3, "SVAGHET", "SkipNav finns men: inga ARIA-labels på ikoner, ingen focus-indicator, ingen dark mode, touch targets ibland <44px.", "WCAG 2.1 AA audit. ARIA-labels. Focus rings. Dark mode. Min 44px touch."),
    ("Emotionell design", 4, "SVAGHET", "Nödbromsen är fantastisk. Men resten av appen är klinisk och transaktionell. Ingen grief support, inga resurser, ingen empati efter dag 7.", "Empatiska microcopy på varje sida. Krisresurser i footer. Mjukare språk."),
    ("Samarbete / multi-user", 2, "SVAGHET", "Invite-länk kopieras till clipboard — det är allt. Ingen realtids-synk, inga kommentarer, ingen uppgiftsfördelning.", "Multi-user med per-delägare view. Kommentarer. Uppgiftstilldelning."),
    ("Data-säkerhet & tillit", 5, "MEDEL", "Supabase med RLS, EU-lagring. Men: lösöre/kostnader i localStorage. Ingen BankID. Ingen tydlig 'datan är sparad'-indikator.", "BankID-inloggning. Flytta ALL data till Supabase. Spara-indikator."),
    ("Telefonsupport", 0, "SAKNAS", "Vi har ingen support alls. Ingen chat, ingen telefon, ingen mail. Alla konkurrenter har åtminstone chat.", "PRIORITET 1: Åtminstone en support-chat. Helst telefon dagtid."),
    ("Jurist-tillgång", 0, "SAKNAS", "Vi har ingen jurist. Aatos har gratis chat-jurist. Lexly/FJ har jurist inkluderad. Detta är vår STÖRSTA svaghet.", "Partner-program med juristbyrå. Eller: AI-driven juridisk FAQ som backup."),
    ("Mobile UX", 7, "STYRKA", "Mobile-first design. BottomNav. Touch-vänligt mest. Men: långa formulär scrolla mycket, modaler täcker skärmen, ingen swipe-nav.", "Dela långa formulär i multi-step. Swipe between sections."),
    ("SEO & Content", 5, "MEDEL", "Landing page finns men thin content. Inga blogg-artiklar, inga guides som rankar i Google. Aatos har 50+ SEO-artiklar.", "Content marketing: 20+ artiklar om dödsbo, arv, bouppteckning."),
    ("Prisvärdhet", 8, "STYRKA", "3 499 kr för ALLT — billigare än alla utom Aatos (1 990). Men Aatos inkluderar jurist-chat, vi inte.", "Överväg lägre pris (2 499?) eller inkludera jurist-chat för att matcha Aatos."),
]

str_fill = PatternFill("solid", fgColor=L_GREEN)
weak_fill = PatternFill("solid", fgColor=L_RED)
med_fill = PatternFill("solid", fgColor=L_YELLOW)
miss_fill = PatternFill("solid", fgColor="F1948A")

for r, (area, score, verdict, honest, action) in enumerate(assessment, 2):
    row_data = [area, score, verdict, honest, action]
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.font = body_sm
        cell.alignment = wrap
        cell.border = thin
        if c == 3:
            if val == "STYRKA": cell.fill = str_fill
            elif val == "SVAGHET": cell.fill = weak_fill
            elif val == "MEDEL": cell.fill = med_fill
            elif val == "SAKNAS": cell.fill = miss_fill
        if c == 2:
            cell.alignment = center
            if score >= 7: cell.fill = str_fill
            elif score >= 5: cell.fill = med_fill
            elif score >= 1: cell.fill = weak_fill
            else: cell.fill = miss_fill

for col, w in [("A",24),("B",10),("C",14),("D",60),("E",55)]:
    ws3.column_dimensions[col].width = w
ws3.auto_filter.ref = f"A1:E{len(assessment)+1}"

# ══════════════════════════════════════════════════════════════
# SHEET 4: PRIORITY ACTION PLAN
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Atgardsplan")

cols4 = ["Prioritet", "Atgard", "Varfor", "Pavekan (1-10)", "Insats", "Kan vi gora det nu?"]
style_header(ws4, cols4)

actions = [
    ("AKUT", "Lägg till support-kanal (chat/mail)", "Vi är ENDA tjänsten utan någon form av support. En sorgande användare som fastnar har INGEN att fråga. Dealbreaker.", 10, "Medel — behöver beslut om chat-lösning", "Delvis — kan sätta upp e-post nu"),
    ("AKUT", "Jurist-partnerskap eller AI-juristchat", "Alla konkurrenter har jurist. Utan detta är vi 'bara en checklista'. Aatos ger gratis juristchat — vi ger NOLL.", 10, "Hög — kräver partneravtal", "NEJ — kräver affärsbeslut"),
    ("AKUT", "Smart dashboard — visa top 3 tasks", "18 länkar utan prioritering = paralysis. Sorgande users behöver 'gör DETTA nu'. Inte 'här är allt du KAN göra'.", 9, "Låg — frontend-ändring", "JA"),
    ("AKUT", "Klickbara telefonnummer (tel:-länkar)", "Nödbroms listar telefonnummer men de är ej klickbara. En sorgande user på mobil ska kunna TAPPA och ringa.", 8, "Mycket låg", "JA"),
    ("HÖG", "BankID-inloggning", "Äldre users (65+) har BankID men klarar inte lösenord+e-post. 8/50 testpersoner flaggade detta.", 8, "Hög — kräver BankID-integration", "NEJ — extern tjänst"),
    ("HÖG", "WCAG 2.1 AA tillgänglighetsaudit", "5/50 personas med funktionsnedsättning kunde ej använda appen. Inga ARIA-labels, dåliga focus indicators.", 7, "Medel", "JA"),
    ("HÖG", "Flytta lösöre + kostnader till Supabase", "Data i localStorage försvinner vid rensning. Users förlorar timmar av arbete. Allvarligt dataförlust-risk.", 8, "Medel", "JA"),
    ("HÖG", "Data-export (PDF/CSV)", "Ingen backup-möjlighet. Om browsern kraschar = allt borta. Ingen kan dela data med jurist.", 7, "Låg", "JA"),
    ("HÖG", "Content marketing — 20+ SEO-artiklar", "Aatos rankar för ALLA dödsbo-sökningar med 50+ artiklar. Vi har noll synlighet i Google.", 8, "Hög — kräver content-produktion", "JA — kan generera"),
    ("MEDEL", "Empatisk microcopy på alla sidor", "Efter nödbromsen blir tonen klinisk. 'Arvskifte' låter som skattedeklaration. Behöver mänskligt språk.", 6, "Låg", "JA"),
    ("MEDEL", "Push-notiser / e-postpåminnelser", "15/50 testpersoner glömde att kolla appen. Deadlines missades. PWA-notiser eller scheduled email.", 7, "Medel", "Delvis — PWA stöder det"),
    ("MEDEL", "Sambo-bodelning i bodelning-modulen", "Sambor är en STOR grupp i Sverige med unik juridisk situation. Bodelning-sidan ignorerar dem helt.", 6, "Låg", "JA"),
    ("MEDEL", ".docx-export för fullmakter", "Users kopierar text → klistrar i Word. Direkt .docx-nedladdning sparar tid och ser mer professionellt ut.", 5, "Låg", "JA"),
    ("MEDEL", "Tooltips för juridiska termer på alla sidor", "Ordlistan finns men users måste navigera dit. Inline tooltips vid varje juridisk term = bättre.", 6, "Medel", "JA"),
    ("MEDEL", "Personnummer-validering med formatering", "Boupptecknings-formuläret tar fri text. Borde validera YYMMDD-XXXX format och Luhn-kontroll.", 5, "Låg", "JA"),
    ("LÅG", "Dark mode", "2/50 personas önskade det. Bra UX men ej kritiskt.", 3, "Medel", "JA"),
    ("LÅG", "Kryptovalutor som tillgångstyp", "2/50 personas. Nischigt men framtidssäkert.", 3, "Låg", "JA"),
    ("LÅG", "Visuellt släktträd", "4/50 personas. Hjälper förstå arvsordning. Men ej kritiskt.", 4, "Hög", "Delvis"),
    ("LÅG", "Flerspråksstöd (engelska)", "3/50 personas. Viktigt för inkludering men stor insats.", 4, "Mycket hög", "NEJ — kräver översättning"),
    ("LÅG", "Digital förrättning via video", "1/50 personas. Juridiskt osäkert om det accepteras.", 2, "Hög", "NEJ — juridisk osäkerhet"),
]

prio_fills = {
    "AKUT": PatternFill("solid", fgColor="F1948A"),
    "HÖG": PatternFill("solid", fgColor="FAD7A0"),
    "MEDEL": PatternFill("solid", fgColor="AED6F1"),
    "LÅG": PatternFill("solid", fgColor=L_GRAY),
}

for r, (prio, action, why, impact, effort, now) in enumerate(actions, 2):
    row_data = [prio, action, why, impact, effort, now]
    for c, val in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        cell.font = body_sm
        cell.alignment = wrap
        cell.border = thin
        if c == 1 and val in prio_fills:
            cell.fill = prio_fills[val]
        if c == 4:
            cell.alignment = center
            if val >= 8: cell.fill = PatternFill("solid", fgColor=L_RED)
            elif val >= 6: cell.fill = PatternFill("solid", fgColor=L_YELLOW)
        if c == 6:
            v = str(val).upper()
            if v.startswith("JA"): cell.fill = PatternFill("solid", fgColor=L_GREEN)
            elif v.startswith("NEJ"): cell.fill = PatternFill("solid", fgColor=L_RED)
            elif "DELVIS" in v: cell.fill = PatternFill("solid", fgColor=L_YELLOW)

for col, w in [("A",10),("B",35),("C",55),("D",10),("E",28),("F",28)]:
    ws4.column_dimensions[col].width = w
ws4.auto_filter.ref = f"A1:F{len(actions)+1}"

# ══════════════════════════════════════════════════════════════
# SHEET 5: COMPETITIVE POSITIONING SUMMARY
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Positionering")

ws5.merge_cells("A1:F1")
ws5.cell(row=1, column=1, value="Dödsboappen — Konkurrenspositionering").font = Font(name="Arial", bold=True, size=16, color=BRAND)

r = 3
for section, items in [
    ("VAD VI GÖR BÄTTRE ÄN ALLA", [
        "Bankbrev auto-genererade från onboarding-data — ingen konkurrent har detta",
        "Kostnadsregistrering per delägare — unikt, förhindrar syskonbråk",
        "Lösöre-inventering med tilldelning och fairness-check — unikt",
        "Avsluta konton-checklista med 34+ poster och kontaktinfo — unikt",
        "Nödbroms dag 1-7 med empatisk steg-för-steg — unikt",
        "Dödsboanmälan smart auto-bedömning mot prisbasbelopp — unikt",
        "Bodelning digital kalkylator — ingen annan digital tjänst har detta",
        "PWA med offline-stöd — enda dödsbo-tjänsten som fungerar utan internet",
        "Ordlista med 37 sökbara juridiska termer — unikt",
        "Tidslinje med auto-beräknade frister — unikt",
    ]),
    ("VAD KONKURRENTERNA GÖR BÄTTRE", [
        "Jurist-tillgång: Aatos har gratis chat-jurist, Lexly/FJ har jurist inkluderad — VI HAR NOLL",
        "Support: ALLA har minst chat/telefon — vi har INGENTING",
        "BankID: Lexly och Familjens Jurist har BankID — vi har lösenord",
        "E-signatur: Lexly har digital signering — vi kräver utskrift",
        "SEO: Aatos har 50+ artiklar som rankar i Google — vi har 0",
        "Trovärdighet: Familjens Jurist har fysiska kontor i hela Sverige",
        "Dokumentkvalitet: Lexly/FJ:s jurister granskar varje dokument — våra PDF:er är ogranskaade",
        "Samarbete: Lexly erbjuder video-möten med alla arvingar — vi har bara en invite-länk",
    ]),
    ("VÅR SWEET SPOT (idealanvändare)", [
        "Digitalt bekväma 25-55-åringar",
        "Enkla till medelkomplexa dödsbon (1-4 delägare, ingen fastighet utomlands)",
        "Vill spara pengar vs jurist (3 499 vs 6 000-103 000)",
        "Vill göra det själv med vägledning — inte outsourca helt",
        "Har tid att hantera processen (ej under extrem tidspress)",
    ]),
    ("VÅR DÖDA VINKEL (vem vi förlorar)", [
        "Äldre (70+) utan teknisk vana — behöver telefonsupport vi ej har",
        "Komplexa dödsbon med företag/utlandstillgångar — vi ger bara disclaimers",
        "Familjer i konflikt — vi har guide men ingen riktig medling",
        "Users som vill ha trygghet av jurist — vi har ingen",
        "Users som hittar oss via Google — vi har ingen SEO-närvaro",
    ]),
]:
    cell = ws5.cell(row=r, column=1, value=section)
    cell.font = Font(name="Arial", bold=True, size=12, color=WHITE)
    cell.fill = hdr_fill
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    for item in items:
        cell = ws5.cell(row=r, column=1, value=f"• {item}")
        cell.font = body
        cell.alignment = wrap
        ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    r += 1

ws5.column_dimensions["A"].width = 120

# ── Save ──
output_path = "konkurrent-analys-och-sjalvbedomning.xlsx"
wb.save(output_path)
print(f"\nRapporten sparad: {output_path}")
print("5 flikar: Konkurrentomdömen | Funktionsjämförelse | Vår självbedömning | Åtgärdsplan | Positionering")
